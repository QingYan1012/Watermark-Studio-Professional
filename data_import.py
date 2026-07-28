# -*- coding: utf-8 -*-
"""
批量数据导入：读取 Excel/CSV，把每一行的字段和一张图片对应起来。

注：这里故意不用 pandas/numpy —— 对于“读一张表格”这种简单需求，pandas+numpy
体积很大（打包后单是这两个库常常就有大几十上百 MB），换成 openpyxl（读 xlsx，
本来就是项目依赖）+ Python 内置 csv 模块（读 csv），打包体积会小很多，
功能上完全够用。旧版 .xls（微软 97-2003 二进制格式）不再支持，遇到时提示
另存为 .xlsx 即可。

第23轮新增：编号+尾部“token 后缀”智能兜底匹配（解决图片带孔号前缀、表格不带
导致精确匹配全 0 的痛点；只在前述匹配都失败时启用，精确结果优先、不被抢占）。

本轮（第24轮）新增：把匹配逻辑收进 match_rows_to_images_detail，额外返回 stats
= {exact, fuzzy, total, mode}，供匹配弹窗“显式说明”智能匹配救场（精确 0 但
fuzzy>0 时弹窗可解释）。旧 match_rows_to_images 收成薄包装，签名/返回值一字
不变 → legacy 的 _count_match / 导出逻辑零影响。

回退：
    set WS_MATCH_FUZZY=0     # 关闭智能兜底，回到“仅精确/标准化匹配”
"""

import os
import re
import csv
import logging
import unicodedata

from openpyxl import load_workbook


_log = logging.getLogger("ws.match")

# 智能兜底总开关（默认开）。关闭后匹配行为与第 1 轮完全一致。
_FUZZY_SUFFIX = os.environ.get("WS_MATCH_FUZZY", "1") != "0"

# token：连续字母串 / 连续数字串 / 连续 CJK 中文串；其余(括号/空格/连字符/点…)全丢弃。
_TOKEN_RE = re.compile(r"[A-Za-z]+|\d+|[一-鿿]+")


def _tokens(s):
    """文件名 -> token 列表。字母小写化；数字/中文原样。分隔符被 findall 自然丢弃。"""
    s = unicodedata.normalize("NFKC", str(s or ""))
    out = []
    for t in _TOKEN_RE.findall(s):
        out.append(t.lower() if t.isascii() else t)
    return out


def _has_digit(tokens):
    return any(t.isdigit() for t in tokens)


def _is_suffix(long_tokens, short_tokens):
    """short_tokens 是否是 long_tokens 的连续后缀。"""
    n = len(short_tokens)
    return n > 0 and len(long_tokens) >= n and long_tokens[-n:] == short_tokens


def load_table(path):
    """
    读取 Excel(.xlsx/.xlsm)/CSV，返回 (columns: list[str], rows: list[dict])。
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        return _load_excel(path)

    if ext == ".xls":
        raise ValueError(
            "暂不支持旧版 .xls（Excel 97-2003）格式，请用 Excel 另存为 .xlsx 后再导入。"
        )

    return _load_csv(path)


def _cell_to_str(v):
    if v is None:
        return ""

    if isinstance(v, float):
        # Excel 里公式算出来的数（比如“上一行止深 + 进尺”这种累加）在文件内部
        # 存的就已经是带二进制误差的浮点数了（63.26+7 存成 70.25999999999999），
        # Excel 显示时按单元格格式圆整过看不出来，但 openpyxl 读到的是原始值。
        # 这里先按 6 位小数圆整消掉浮点噪声，再去掉多余的尾随 0，
        # 正常需要 2~3 位小数的深度/坐标类数值不受影响。
        r = round(v, 6)
        if r == int(r):
            return str(int(r))
        return f"{r:.6f}".rstrip("0").rstrip(".")

    return str(v)


def _load_excel(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            return [], []

        columns = [
            (_cell_to_str(c) or f"列{i + 1}")
            for i, c in enumerate(header)
        ]

        rows = []

        for r in rows_iter:
            if r is None or all(v is None for v in r):
                continue

            row = {}
            for i, col in enumerate(columns):
                row[col] = _cell_to_str(r[i]) if i < len(r) else ""

            rows.append(row)

        return columns, rows

    finally:
        wb.close()


def _load_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        raw_rows = list(csv.reader(f))

    if not raw_rows:
        return [], []

    columns = [
        (_cell_to_str(c) or f"列{i + 1}")
        for i, c in enumerate(raw_rows[0])
    ]

    rows = []

    for r in raw_rows[1:]:
        if not r or all((v is None or v == "") for v in r):
            continue

        row = {}
        for i, col in enumerate(columns):
            row[col] = r[i] if i < len(r) else ""

        rows.append(row)

    return columns, rows


def _normalize_filename(s):
    """
    把文件名“标准化”用于兜底匹配：

    - 全角转半角（Excel 手打或输入法容易把 “-” 打成 “－”、数字打成全角）
    - 大小写不敏感
    - 掐头去尾空格
    - 内部多个空格合并成一个

    只在精确匹配失败时才会用这个兜底，避免“文件名”列本身就有细微差异（大小写、
    全角符号）时被误判成“没有匹配上”。
    """
    s = unicodedata.normalize("NFKC", str(s))
    s = s.strip().casefold()
    s = " ".join(s.split())
    return s


def match_rows_to_images_detail(rows, image_paths, filename_column=None):
    """
    把表格行和图片一一对应，并返回匹配统计。

    返回 (mapping, stats)：
        mapping : dict, image_path -> row(dict)（没匹配到的图片得到空 dict {}）
        stats   : dict, 含
            exact : 精确 + 标准化匹配命中的图数
            fuzzy : 编号+尾部 token 后缀智能兜底额外命中的图数
            total : 最终匹配上的图数（= exact + fuzzy；顺序模式 = 顺序对应非空数）
            mode  : "filename" 或 "order"

    - 指定 filename_column 时：精确 → 标准化 →【第23轮】token 后缀兜底，逐级回退；
      精确结果优先，绝不被兜底抢占；双向唯一（一行一图）。
    - 不指定时：按顺序对应（第 i 行 -> 第 i 张图），此时 exact=fuzzy=0。
    """
    result = {}

    if filename_column:
        lookup_full = {}        # 完整文件名（含扩展名）-> row
        lookup_stem = {}        # 不含扩展名的文件名 -> row
        lookup_full_norm = {}   # 标准化后的完整文件名 -> row（兜底）
        lookup_stem_norm = {}   # 标准化后的不含扩展名文件名 -> row（兜底）

        for row in rows:
            key = str(row.get(filename_column, "")).strip()
            if not key:
                continue

            stem = os.path.splitext(key)[0]

            lookup_full.setdefault(key, row)
            lookup_stem.setdefault(stem, row)

            lookup_full_norm.setdefault(_normalize_filename(key), row)
            lookup_stem_norm.setdefault(_normalize_filename(stem), row)

        # ---- 阶段 A：精确 + 标准化匹配，记录被占用的行 + 命中数 exact_n ----
        occupied = set()
        exact_n = 0
        fuzzy_n = 0   # 在此初始化，保证 _FUZZY_SUFFIX=False 时也有定义

        for p in image_paths:
            base = os.path.basename(p)
            stem = os.path.splitext(base)[0]

            row = (
                lookup_full.get(base)
                or lookup_stem.get(stem)
                or lookup_stem.get(base)
                or lookup_full_norm.get(_normalize_filename(base))
                or lookup_stem_norm.get(_normalize_filename(stem))
                or lookup_stem_norm.get(_normalize_filename(base))
            )

            if row:
                result[p] = dict(row)
                occupied.add(id(row))
                exact_n += 1
            else:
                result[p] = {}

        # ---- 阶段 B：编号+尾部 token 后缀智能兜底 ----
        # 只补阶段 A 没配上的图，只用阶段 A 没占用的行；精确结果绝不被抢占。
        if _FUZZY_SUFFIX:
            free = []
            for row in rows:
                if id(row) in occupied:
                    continue
                key = str(row.get(filename_column, "")).strip()
                if not key:
                    continue
                rt = _tokens(os.path.splitext(key)[0])
                # 太宽泛的行(无数字/过短)不放行，避免“纯描述串”匹配所有图
                if len(rt) >= 2 and _has_digit(rt):
                    free.append([row, rt])

            for p in image_paths:
                if result[p]:
                    continue

                it = _tokens(os.path.splitext(os.path.basename(p))[0] or os.path.basename(p))
                if not it:
                    continue

                best_k = -1
                best_row = None
                best_len = -1

                for k, (_row, rt) in enumerate(free):
                    if _is_suffix(it, rt) and len(rt) > best_len:
                        best_len = len(rt)
                        best_k = k
                        best_row = _row

                if best_row is not None:
                    result[p] = dict(best_row)
                    occupied.add(id(best_row))
                    fuzzy_n += 1
                    try:
                        _log.debug(
                            "fuzzy suffix match: img=%s <- row tokens=%s",
                            os.path.basename(p),
                            free[best_k][1],
                        )
                    except Exception:
                        pass
                    free.pop(best_k)   # 该行已占用，双向唯一

            if fuzzy_n:
                try:
                    _log.info("智能兜底(编号+尾部)额外匹配 %d 张", fuzzy_n)
                except Exception:
                    pass

        mode = "filename"

    else:
        for i, p in enumerate(image_paths):
            result[p] = dict(rows[i]) if i < len(rows) else {}

        exact_n = 0
        fuzzy_n = 0
        mode = "order"

    total = sum(1 for d in result.values() if d)
    stats = {"exact": exact_n, "fuzzy": fuzzy_n, "total": total, "mode": mode}

    try:
        _log.debug(
            "match stats: mode=%s exact=%d fuzzy=%d total=%d",
            mode, exact_n, fuzzy_n, total,
        )
    except Exception:
        pass

    return result, stats


def match_rows_to_images(rows, image_paths, filename_column=None):
    """
    旧接口（薄包装）：签名与返回值与第 1 轮完全一致，返回 mapping(dict)。

    legacy 的 _count_match / 导出逻辑仍调本函数，行为零变化。
    需要 exact/fuzzy 拆分的新代码请调 match_rows_to_images_detail。
    """
    return match_rows_to_images_detail(rows, image_paths, filename_column)[0]


def diagnose_filename_mismatch(rows, image_paths, filename_column, sample_n=3):
    """
    在按文件名匹配后发现匹配数很低/为零时，用来给用户一个直观对比：

    - 表格里实际读到的文件名长什么样
    - 图片实际的文件名又长什么样

    方便一眼看出是“多了个空格”“大小写不同”“前面还带了个文件夹路径”这类问题。

    返回 (excel_samples, image_samples)，各最多 sample_n 条。
    """
    excel_samples = []

    for row in rows:
        v = str(row.get(filename_column, "")).strip()
        if v:
            excel_samples.append(v)
        if len(excel_samples) >= sample_n:
            break

    image_samples = [
        os.path.basename(p)
        for p in image_paths[:sample_n]
    ]

    return excel_samples, image_samples